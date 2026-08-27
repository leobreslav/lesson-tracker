"""
Учебный план в xlsx: чтение книги и запись её же.

Формат тот же самый, что у CSV, — три столбца `id, Тема, Урок`,
одна строка это один урок, — и разбор структуры общий: здесь только чтение
ячеек, а решает всё `services.parse_plan_rows`. Ради этого модуль и заведён
отдельно: единственное место, знающее про openpyxl, как `files/storage.py`
единственное место, знающее про boto3.

Зачем xlsx рядом с CSV. Учителя работают в Excel и Google Sheets, а CSV
каждый раз приносит одно и то же: кодировку, разделитель и кавычки в
названиях. В книге ничего этого нет — там сразу ячейки.
"""

from __future__ import annotations

import datetime as dt
import io
from typing import Iterable, NamedTuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Protection
from openpyxl.utils import get_column_letter

from .services import (
    CSV_HEADER,
    CSV_HEADER_DATED,
    Branch,
    PlanImportError,
    plan_date_cell,
)

SHEET_TITLE = "План"

# ширина столбца: минимум, запас на символ и потолок. Потолок нужен: одна
# заметка в пол-экрана иначе растянула бы столбец на всю ширину листа.
# Четвёртый — дата, и она всегда одной длины, поэтому минимум равен потолку
COLUMN_SIZES = ((5, 8), (24, 50), (32, 70), (12, 12))

TEXT_FORMAT = "@"


class ReadWorkbook(NamedTuple):
    """Строки первого листа и то, что стоит сказать о самой книге."""

    rows: list
    sheet: str
    sheets_ignored: int


def cell_text(value) -> str:
    """
    Ячейка в строку — так, как её видел человек.

    Excel хранит в ячейке не то, что в ней написано: id уезжает числом
    (412 становится `412.0`), а название вроде «10.09» — датой. Считать
    такую ячейку испорченной нельзя, файл-то на вид правильный, поэтому
    приводим обратно к тексту и разбираем дальше как обычно.
    """
    if value is None:
        return ""

    if isinstance(value, bool):
        return "ИСТИНА" if value else "ЛОЖЬ"

    if isinstance(value, float):
        # 412.0 — это id 412, а не «412.0»
        return str(int(value)) if value.is_integer() else repr(value)

    if isinstance(value, dt.datetime):
        return (
            value.date().isoformat()
            if value.time() == dt.time()
            else value.isoformat(sep=" ")
        )

    if isinstance(value, (dt.date, dt.time)):
        return value.isoformat()

    return str(value)


def read_plan_xlsx(data: bytes, *, filename: str = "") -> ReadWorkbook:
    """
    Книга → строки ячеек. Читается первый лист, остальные называются в ответе.

    Старый `.xls` — другой формат, не zip с XML внутри, и openpyxl его не
    открывает. Сказать об этом прямо дешевле, чем тащить вторую библиотеку
    ради формата, который Excel не пишет по умолчанию с 2007 года.
    """
    if filename.lower().endswith(".xls"):
        raise PlanImportError(
            "The old «.xls» format is not supported — save the file as «.xlsx»."
        )

    try:
        book = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        # openpyxl поднимает своё на битом zip, KeyError на чужом zip и
        # ещё несколько разных — для нас это один случай: не книга
        raise PlanImportError(
            "The file is not an Excel workbook — «.xlsx» is expected."
        )

    try:
        sheet = book.worksheets[0]
        rows = [
            [cell_text(value) for value in raw]
            for raw in sheet.iter_rows(values_only=True)
        ]
        return ReadWorkbook(rows, sheet.title, len(book.worksheets) - 1)
    finally:
        book.close()


def build_plan_xlsx(tree: Iterable[Branch], dates=None) -> bytes:
    """
    План книгой — тем же форматом, который читает импорт.

    Оформление здесь не украшение, а защита от того, чем xlsx портят:
    текстовый формат ячеек не даёт Excel превратить «10.09» в дату,
    закреплённая шапка не уезжает на длинном плане, а лист защищён с
    разблокированными ячейками — кроме столбцов, которые правит не человек.
    Пароля нет: снять защиту легко, задача только в том, чтобы их не стёрли
    не глядя.

    `dates` — раскладка, `pk строки плана → дата`; с ней появляется четвёртый
    столбец. Он заперт наравне с id, и по той же причине, только сильнее:
    id правкой хотя бы можно испортить план, а дата приедет обратно ничем —
    её место в расписании, и импорт этот столбец отбрасывает.
    """
    header = CSV_HEADER_DATED if dates is not None else CSV_HEADER

    book = Workbook()
    sheet = book.active
    sheet.title = SHEET_TITLE

    widths = [len(title) for title in header]
    line = 0

    def put(values, *, header=False):
        nonlocal line
        line += 1
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=line, column=index, value=value)
            cell.number_format = TEXT_FORMAT
            cell.font = Font(bold=header)
            cell.alignment = Alignment(
                wrap_text=index in (2, 3) and not header,
                vertical="center" if header else "top",
            )
            # заперты id и дата: их правит не человек, а остальное для того и
            # выгружено, чтобы правили
            cell.protection = Protection(locked=index in (1, 4))
            widths[index - 1] = max(widths[index - 1], len(str(value)))

    put(header, header=True)

    def row(node, theme):
        cells = [node.pk, theme, node.title]
        if dates is not None:
            cells.append(plan_date_cell(dates, node))
        return cells

    for branch in tree:
        if branch.node.is_section:
            for child in branch.children:
                put(row(child, branch.node.title))
        else:
            put(row(branch.node, ""))

    for index, (minimum, cap) in enumerate(COLUMN_SIZES[:len(header)], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = min(
            max(minimum, widths[index - 1] + 2), cap
        )

    sheet.freeze_panes = "A2"
    sheet.protection.sheet = True

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()
