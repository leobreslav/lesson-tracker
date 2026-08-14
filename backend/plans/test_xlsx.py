"""
Учебный план в xlsx: чтение книги, запись книги, круговой прогон.

Формат тот же, что у CSV, и разбор структуры общий, поэтому здесь не
повторяется вся матрица отказов — она в `test_csv.py`. Проверяется то, чего
у CSV нет вовсе: как Excel хранит значения (id числом, название датой),
оформление листа и то, что защита столбца id не мешает импортировать
собственную же выгрузку.
"""

import datetime as dt
import io

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import SimpleTestCase
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from . import xlsx
from .models import PlanNode
from .tests import PlanTestCase

XLSX_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
HEADER = ("id", "Тема", "Урок", "Заметка")


def book_bytes(rows, *, sheets=(), title="План"):
    """Книга из строк — то, что мог бы сохранить Excel или Google Sheets."""
    book = Workbook()
    sheet = book.active
    sheet.title = title
    for row in rows:
        sheet.append(list(row))
    for name in sheets:
        book.create_sheet(name)

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


class CellTextTests(SimpleTestCase):
    """Ячейка в строку — так, как её видел человек."""

    def test_a_whole_float_is_an_id(self):
        # Excel хранит числа как float: 412 приезжает как 412.0
        self.assertEqual(xlsx.cell_text(412.0), "412")

    def test_a_number_stays_a_number(self):
        self.assertEqual(xlsx.cell_text(412), "412")

    def test_an_empty_cell_is_an_empty_string(self):
        self.assertEqual(xlsx.cell_text(None), "")

    def test_a_date_becomes_a_readable_date(self):
        """«10.09» Excel превращает в дату — название всё равно должно доехать."""
        self.assertEqual(xlsx.cell_text(dt.datetime(2026, 9, 10)), "2026-09-10")

    def test_a_fractional_number_keeps_its_dot(self):
        self.assertEqual(xlsx.cell_text(1.5), "1.5")


class ReadTests(SimpleTestCase):
    def test_the_first_sheet_is_read_and_the_others_are_named(self):
        data = book_bytes(
            [HEADER, ["", "Векторы", "Понятие", ""]], sheets=("Черновик", "Заметки")
        )

        book = xlsx.read_plan_xlsx(data)

        self.assertEqual(book.sheet, "План")
        self.assertEqual(book.sheets_ignored, 2)
        self.assertEqual(book.rows[1], ["", "Векторы", "Понятие", ""])

    def test_the_old_xls_format_is_refused_by_name(self):
        with self.assertRaises(xlsx.PlanImportError) as caught:
            xlsx.read_plan_xlsx(b"\xd0\xcf\x11\xe0", filename="план.xls")

        self.assertIn(".xlsx", str(caught.exception))

    def test_a_file_that_is_not_a_workbook_is_refused(self):
        with self.assertRaises(xlsx.PlanImportError):
            xlsx.read_plan_xlsx(b"%PDF-1.7\n%\xc7\xec\x8f\xa2", filename="план.xlsx")


class ExportTests(PlanTestCase):
    def export(self, course=None):
        return self.client.get(
            reverse("plannode-export-xlsx"), {"course": (course or self.course).pk}
        )

    def sheet(self, response):
        return load_workbook(io.BytesIO(response.content)).worksheets[0]

    def test_export_writes_one_row_per_lesson(self):
        self.build_sample()
        sine = self.node("Синус суммы")

        sheet = self.sheet(self.export())

        self.assertEqual([cell.value for cell in sheet[1]], list(HEADER))
        self.assertEqual(
            [xlsx.cell_text(cell.value) for cell in sheet[2]],
            [str(sine.pk), "Тригонометрия", "Синус суммы", ""],
        )

    def test_the_header_is_bold_and_frozen(self):
        self.build_sample()

        sheet = self.sheet(self.export())

        self.assertTrue(sheet["A1"].font.bold)
        self.assertEqual(sheet.freeze_panes, "A2")

    def test_the_id_column_is_the_only_locked_one(self):
        """
        Пароля нет: снять защиту легко. Задача в том, чтобы id не стёрли,
        не заметив, — а править названия ради этого мешать нельзя.
        """
        self.build_sample()

        sheet = self.sheet(self.export())

        self.assertTrue(sheet.protection.sheet)
        self.assertTrue(sheet["A2"].protection.locked)
        self.assertFalse(sheet["C2"].protection.locked)

    def test_cells_are_text_so_excel_keeps_its_hands_off(self):
        self.build_sample()

        sheet = self.sheet(self.export())

        self.assertEqual(sheet["C2"].number_format, "@")

    def test_the_theme_and_lesson_columns_wrap(self):
        self.build_sample()

        sheet = self.sheet(self.export())

        self.assertTrue(sheet["B2"].alignment.wrap_text)
        self.assertTrue(sheet["C2"].alignment.wrap_text)
        self.assertFalse(sheet["A2"].alignment.wrap_text)

    def test_columns_are_sized_by_content_but_capped(self):
        self.add("У" * 190, position=0)

        sheet = self.sheet(self.export())

        self.assertLess(sheet.column_dimensions["A"].width, 10)
        self.assertLessEqual(sheet.column_dimensions["C"].width, 60)
        self.assertGreater(sheet.column_dimensions["C"].width, 20)

    def test_the_filename_carries_the_course_and_the_date(self):
        response = self.export()

        disposition = response["Content-Disposition"]
        self.assertIn(".xlsx", disposition)
        self.assertIn("%D0%BF%D0%BB%D0%B0%D0%BD", disposition)  # «план»

    def test_cannot_export_another_users_course(self):
        self.assertEqual(self.export(self.alien_class).status_code, 404)

    def test_requires_authentication(self):
        self.client.credentials()

        self.assertEqual(self.export().status_code, 401)


class ImportTests(PlanTestCase):
    def upload(self, data, mode="replace", name="план.xlsx", course=None):
        return self.client.post(
            f"{reverse('plannode-import-xlsx')}?course={(course or self.course).pk}",
            {
                "file": SimpleUploadedFile(name, data, content_type=XLSX_TYPE),
                "mode": mode,
            },
            format="multipart",
        )

    def test_import_builds_the_tree(self):
        data = book_bytes(
            [
                HEADER,
                ["", "Векторы", "Понятие вектора", "повторить"],
                ["", "Векторы", "Сложение векторов", ""],
                ["", "", "Повторение", ""],
            ]
        )

        response = self.upload(data)

        self.assertEqual(response.status_code, 200, response.content)
        self.assertEqual(
            self.structure(),
            ["Векторы", "  Понятие вектора", "  Сложение векторов", "Повторение"],
        )
        self.assertEqual(response.json()["sheet"], "План")
        self.assertEqual(response.json()["sheets_ignored"], 0)

    def test_a_second_sheet_is_ignored_and_said_so(self):
        data = book_bytes([HEADER, ["", "Векторы", "Понятие", ""]], sheets=("Ещё",))

        response = self.upload(data)

        self.assertEqual(response.json()["sheets_ignored"], 1)
        self.assertEqual(self.structure(), ["Векторы", "  Понятие"])

    def test_an_id_saved_as_a_float_still_matches_the_row(self):
        """Excel хранит числа как float, и 412 приезжает как 412.0."""
        section = self.add("Векторы", position=0, is_section=True)
        lesson = self.add("Понятие", parent=section, position=0)
        lesson.body = "текст урока"
        lesson.save(update_fields=["body"])

        data = book_bytes([HEADER, [float(lesson.pk), "Векторы", "Понятие", ""]])

        response = self.upload(data, mode="sync")

        self.assertEqual(response.status_code, 200, response.content)
        self.assertEqual(response.json()["created"], 0)
        lesson.refresh_from_db()
        self.assertEqual(lesson.body, "текст урока")

    def test_empty_rows_at_the_end_are_ignored(self):
        data = book_bytes(
            [HEADER, ["", "Векторы", "Понятие", ""], [None, None, None, None], []]
        )

        response = self.upload(data)

        self.assertEqual(response.status_code, 200, response.content)
        self.assertEqual(self.structure(), ["Векторы", "  Понятие"])

    def test_commas_and_quotes_need_no_escaping_here(self):
        """Ради этого формат и добавлен: в ячейке нет ни кавычек, ни запятых."""
        data = book_bytes(
            [
                HEADER,
                ["", 'Дроби, обыкновенные и "десятичные"', "Сложение; вычитание", "№ 12, 13"],
            ]
        )

        response = self.upload(data)

        self.assertEqual(response.status_code, 200, response.content)
        self.assertEqual(
            self.titles_with_notes(),
            [
                (None, 'Дроби, обыкновенные и "десятичные"', "", True),
                (
                    'Дроби, обыкновенные и "десятичные"',
                    "Сложение; вычитание",
                    "№ 12, 13",
                    False,
                ),
            ],
        )

    def test_the_same_refusals_as_csv(self):
        cases = {
            "csv_header_invalid": [["Тема", "Урок", "Заметка"], ["Векторы", "", ""]],
            "csv_section_row": [HEADER, ["", "Векторы", "", ""]],
            "csv_bad_id": [HEADER, ["абв", "Векторы", "Понятие", ""]],
            "csv_row_too_long": [HEADER, ["", "Векторы", "о" * 201, ""]],
        }

        for code, rows in cases.items():
            with self.subTest(code=code):
                response = self.upload(book_bytes(rows))

                self.assertEqual(response.status_code, 400, response.content)
                self.assertEqual(response.json()["code"], code)
                self.assertFalse(
                    PlanNode.objects.filter(course=self.course).exists()
                )

    def test_the_old_xls_format_is_refused_clearly(self):
        response = self.upload(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1", name="план.xls")

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()["code"], "file_not_xlsx")
        self.assertIn(".xlsx", response.json()["detail"])

    def test_a_renamed_pdf_is_refused(self):
        response = self.upload(b"%PDF-1.7\n%\xc7\xec\x8f\xa2\n1 0 obj\n", name="план.xlsx")

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()["code"], "file_not_xlsx")

    def test_mode_is_required(self):
        data = book_bytes([HEADER, ["", "Векторы", "Понятие", ""]])

        response = self.client.post(
            f"{reverse('plannode-import-xlsx')}?course={self.course.pk}",
            {"file": SimpleUploadedFile("план.xlsx", data, content_type=XLSX_TYPE)},
            format="multipart",
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()["code"], "mode_required")

    def test_cannot_import_into_another_users_course(self):
        data = book_bytes([HEADER, ["", "Векторы", "Понятие", ""]])

        response = self.upload(data, course=self.alien_class)

        self.assertEqual(response.status_code, 404)


class RoundTripTests(PlanTestCase):
    """Экспорт → импорт: ни одной новой строки и ни одной удалённой."""

    def setUp(self):
        super().setUp()
        self.section = self.add('Дроби, "обыкновенные"; и десятичные',
                                position=0, is_section=True)
        self.first = self.add("Сложение, вычитание", parent=self.section, position=0)
        self.second = self.add("У" * 190, parent=self.section, position=1)
        self.loose = self.add("Повторение", position=1)

        self.first.body = "Пусть $a^2 + b^2 = c^2$."
        self.first.note = "№ 12, 13"
        self.first.save()

    def export(self):
        return self.client.get(
            reverse("plannode-export-xlsx"), {"course": self.course.pk}
        ).content

    def send(self, data, mode="sync"):
        return self.client.post(
            f"{reverse('plannode-import-xlsx')}?course={self.course.pk}",
            {
                "file": SimpleUploadedFile("план.xlsx", data, content_type=XLSX_TYPE),
                "mode": mode,
            },
            format="multipart",
        )

    def test_export_then_sync_changes_nothing(self):
        before = self.titles_with_notes()

        response = self.send(self.export())

        self.assertEqual(response.status_code, 200, response.content)
        self.assertEqual(response.json()["created"], 0)
        self.assertEqual(response.json()["deleted"], 0)
        self.assertEqual(self.titles_with_notes(), before)
        self.first.refresh_from_db()
        self.assertEqual(self.first.body, "Пусть $a^2 + b^2 = c^2$.")

    def test_the_locked_id_column_does_not_get_in_the_way(self):
        """Лист выгружается защищённым — и своя же выгрузка читается обратно."""
        sheet = load_workbook(io.BytesIO(self.export())).worksheets[0]
        self.assertTrue(sheet.protection.sheet)

        self.assertEqual(self.send(self.export()).status_code, 200)

    def test_a_file_saved_again_by_a_spreadsheet_still_imports(self):
        """
        Google Sheets: открыли выгрузку, сохранили обратно.

        При этом теряется оформление, id становится числом, а пустые ячейки
        превращаются в None — на разбор ничего из этого влиять не должно.
        """
        sheet = load_workbook(io.BytesIO(self.export())).worksheets[0]
        head, *body = list(sheet.iter_rows())
        rows = [[cell.value for cell in head]]
        rows += [
            [
                # id снова стал числом: так его отдаёт любая таблица
                float(cell.value) if index == 0 and cell.value else cell.value
                for index, cell in enumerate(row)
            ]
            for row in body
        ]
        before = self.titles_with_notes()

        response = self.send(book_bytes(rows, title="Sheet1"))

        self.assertEqual(response.status_code, 200, response.content)
        self.assertEqual(response.json(), {
            "created": 0, "updated": 4, "deleted": 0,
            "sheet": "Sheet1", "sheets_ignored": 0,
        })
        self.assertEqual(self.titles_with_notes(), before)


class PreviewTests(PlanTestCase):
    def preview(self, data, mode="sync"):
        return self.client.post(
            f"{reverse('plannode-import-preview-xlsx')}?course={self.course.pk}",
            {
                "file": SimpleUploadedFile("план.xlsx", data, content_type=XLSX_TYPE),
                "mode": mode,
            },
            format="multipart",
        )

    def test_preview_says_how_the_book_was_read(self):
        """Клиент xlsx не разбирает — «как прочитано» он узнаёт отсюда."""
        self.add("Старый урок", position=0)
        data = book_bytes(
            [HEADER, ["", "Векторы", "Понятие", ""], ["", "Векторы", "Сложение", ""]],
            sheets=("Черновик",),
        )

        body = self.preview(data, mode="replace").json()

        self.assertEqual(body["rows"], 2)
        self.assertEqual(body["lessons"], 2)
        self.assertEqual(body["sections"], 1)
        self.assertEqual(body["delete"], 1)
        self.assertEqual(body["sheet"], "План")
        self.assertEqual(body["sheets_ignored"], 1)
        self.assertFalse(body["syncable"])

    def test_preview_changes_nothing(self):
        self.add("Старый урок", position=0)

        self.preview(book_bytes([HEADER, ["", "Векторы", "Понятие", ""]]),
                     mode="replace")

        self.assertEqual(self.structure(), ["Старый урок"])

    def test_a_file_that_is_not_a_workbook_is_reported_not_refused(self):
        """
        Предпросмотр не отвечает 400 на выбранный файл.

        Иначе единственным следом отказа была бы ошибка в консоли браузера:
        диалог показывает ошибки из списка, а не разбирает коды HTTP.
        """
        body = self.preview(b"%PDF-1.7\n%\xc7\xec\x8f\xa2\n", mode="replace")

        self.assertEqual(body.status_code, 200)
        self.assertEqual(
            [error["code"] for error in body.json()["errors"]], ["file_not_xlsx"]
        )
